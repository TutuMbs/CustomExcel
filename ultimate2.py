import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Color
from openpyxl.styles import Color


def ajusta_fontes_aligment_celulas(range_colunas: str, workbook, cor_fill=None, tipo_borda=None, tipo_alignment=None,
                                   bold_fonte=None, cell_fonte=None):
    for row in workbook[range_colunas]:
        for cell in row:
            if cor_fill:
                cell.fill = cor_fill
            if tipo_borda:
                cell.border = tipo_borda
            if tipo_alignment:
                cell.alignment = tipo_alignment
            if bold_fonte:
                cell.font = bold_fonte
            if cell_fonte:
                cell.font = cell_fonte


workbook = openpyxl.Workbook()
worksheet = workbook.active


worksheet['A1'] = 'Data Base'
worksheet['B1'] = '31/12/2022'
worksheet['G1'] = ' Seguro - SAG Dez-2023'


worksheet['G2'] = 'Nº  Grupo'
worksheet['H2'] = 'Meses para a Última Assembleia'
worksheet['I2'] = 'Cotas_Ativas_Cont'
worksheet['J2'] = 'Cotas_Ativas_Cont '
worksheet['K2'] = 'Cotas_Ativas_Cont C_Pagto_ Parcial'
worksheet['L2'] = 'Cotas_Ativas_Cont C_Bem_Pendente'
worksheet['M2'] = 'Cotas Ativas a Contemplar'
worksheet['N2'] = 'Ct_Excluidas já Devolvidas '
worksheet['O2'] = 'Ct_Excluidas a Contemplar'
worksheet['P2'] = 'Ct_Excluidas Cont Pend Devolução'
worksheet['Q2'] = 'Ct_Excluidas Cont com Devol Parcial'
worksheet['R2'] = 'Ct_Excluidas_não cont Com Devol total'
worksheet['S2'] = 'Ct_Excluidas Ncont Com Devol Parcial'
worksheet['T2'] = 'Ct Ativas Não Cont em_atraso [1 a 2] parcelas'
worksheet['U2'] = 'Ct Ativas Não Cont em_atraso [ 3 ] parcelas'
worksheet['V2'] = 'Ct Ativas Não Cont em_atraso [ + 3 ] parcelas'
worksheet['W2'] = 'Ct_Ativas_Cont em_Atraso_[1 a 3] Parcelas'
worksheet['X2'] = 'Ct_Ativas_Cont em_Atraso_[+ 3]_parcelas'
worksheet['Y2'] = 'Ct_Ativa_Cont Inad.[+3]_Com_bem'
worksheet['Z2'] = 'Ct Ativa Cont Inad.[+3]_Sem_bem'
worksheet['AA2'] = 'Ct_At_Cont_S_Bem Inad.[+3]_Ajuzadas'
worksheet['AB2'] = 'Ct_At_Cont_C_Bem Inad.[+3]_Ajuzadas'
worksheet['AC2'] = 'Ct_At_Cont_C_Bem Inad.[+3]_Cob Amigavel'
worksheet['AD2'] = 'Ct_At_Cont_Inad.[+3] Ajui_S/bem Retomado'
worksheet['AE2'] = 'Ct_At_Cont_ Inad.[+3] Ajui_C/bem Retomado'

worksheet['G3'] = 'I'
worksheet['H3'] = 'II'
worksheet['I3'] = 'III'
worksheet['J3'] = 'III.1'
worksheet['K3'] = 'III.2'
worksheet['L3'] = 'IV'
worksheet['M3'] = 'V'
worksheet['N3'] = 'VI'
worksheet['O3'] = 'VII'
worksheet['P3'] = 'VIII'
worksheet['Q3'] = 'VIII.1'
worksheet['R3'] = 'VIII.2'
worksheet['S3'] = 'IX'
worksheet['T3'] = 'IX.1'
worksheet['U3'] = 'IX.2'
worksheet['V3'] = 'X'
worksheet['W3'] = 'XI'
worksheet['X3'] = 'XII'
worksheet['Y3'] = 'XIII'
worksheet['Z3'] = 'XIV'
worksheet['AA3'] = 'XV'
worksheet['AB3'] = 'XV.1'
worksheet['AC3'] = 'XVI'
worksheet['AD3'] = 'XVII'
worksheet['AE3'] = 'XVIII'

worksheet['G4'] = 'IMOVEL'
worksheet['G5'] = 'AUTO'
worksheet['G6'] = 'PESADO'
worksheet['G7'] = 'OUTROS'

worksheet['G9'] = 'RESUMO'

# Salve o arquivo do Excel
workbook.save('C.xlsx')


# Criando um arquivo Excel e adicionando uma planilha
wb = load_workbook('C.xlsx')
ws = wb.active
wa = wb.active
wd = wb.active

borda = Border(left=Side(border_style='thin'),
               right=Side(border_style='thin'),
               top=Side(border_style='thin'),
               bottom=Side(border_style='thin'))

# gradiente

vermelho = PatternFill(start_color='FF0000',
                       end_color='FF0000', fill_type='solid')
azulclarinho = PatternFill(start_color='BDD7EE',
                           end_color='BDD7EE', fill_type='solid')
verdeclaro = PatternFill(start_color='E2EFDA',
                         end_color='E2EFDA', fill_type='solid')
amareloclaro = PatternFill(start_color='FFF2CC',
                           end_color='FFF2CC', fill_type='solid')
cinzaescuro = PatternFill(start_color='AEAAAA',
                          end_color='AEAAAA', fill_type='solid')
cinza = PatternFill(start_color='BFBFBF',
                    end_color='BFBFBF', fill_type='solid')
cinzaclaro = PatternFill(start_color='D0CECE',
                         end_color='D0CECE', fill_type='solid')
azul = PatternFill(start_color='8EA9DB',
                   end_color='8EA9DB', fill_type='solid')


fontebranca = Font(color="FFFFFF")
fundoazul = PatternFill(start_color='4472C4',
                        end_color='4472C4', fill_type='solid')

ajusta_fontes_aligment_celulas(
    range_colunas='G1:AH1', workbook=ws, cor_fill=fundoazul, cell_fonte=fontebranca)

ajusta_fontes_aligment_celulas(
    range_colunas='G9:AH9', workbook=ws, cor_fill=fundoazul, cell_fonte=fontebranca)

grau = Alignment(textRotation=45)
ajusta_fontes_aligment_celulas(
    range_colunas='AC2:AD2', workbook=wa, cor_fill=azul, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='G2:G2', workbook=wa, cor_fill=vermelho, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='H2:H2', workbook=wa, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='I2:L2', workbook=wa, cor_fill=azulclarinho, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='M2:R2', workbook=wa, cor_fill=verdeclaro, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='V2:W2', workbook=wa, cor_fill=cinzaescuro, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='S2:U2', workbook=wa, cor_fill=amareloclaro, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='X2:Y2', workbook=wa, cor_fill=cinza, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='Z2:AB2', workbook=wa, cor_fill=cinzaclaro, tipo_alignment=grau)

ajusta_fontes_aligment_celulas(
    range_colunas='AC2:AE2', workbook=wa, cor_fill=azul, tipo_alignment=grau)


grau0 = Alignment(textRotation=0)
azulclaro = PatternFill(start_color='DDEBF7',
                        end_color='DDEBF7', fill_type='solid')
bold_font = Font(bold=True)
Fonte = Font(size=18)

ajusta_fontes_aligment_celulas(range_colunas='G3:AE3', workbook=wa,
                               cor_fill=azulclaro, tipo_borda=borda, tipo_alignment=grau0)


ajusta_fontes_aligment_celulas(
    range_colunas='G3:AE3', workbook=wa, bold_fonte=bold_font, cell_fonte=Fonte)

ajusta_fontes_aligment_celulas(
    range_colunas='G4:AE4', workbook=ws, cor_fill=fundoazul, cell_fonte=fontebranca)

ajusta_fontes_aligment_celulas(
    range_colunas='G5:AE5', workbook=ws, cor_fill=fundoazul, cell_fonte=fontebranca)

ajusta_fontes_aligment_celulas(
    range_colunas='G6:AE6', workbook=ws, cor_fill=fundoazul, cell_fonte=fontebranca)

ajusta_fontes_aligment_celulas(
    range_colunas='G7:AE7', workbook=ws, cor_fill=fundoazul, cell_fonte=fontebranca)

# Salvando o arquivo Excel
wb.save('C.xlsx')
